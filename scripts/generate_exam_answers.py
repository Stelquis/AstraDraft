#!/usr/bin/env python3
"""AstraDraft 最终考核答案生成脚本
读取考核试题 -> 加载CAD图纸 -> 逐题回答 -> 输出答案
"""
import csv
import json
import logging
import os
import sys
import openpyxl

# 添加项目根目录到PATH
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from agent.core import DeepAstraDraft as Agent
from backend.config import AgentConfig

logging.basicConfig(level=logging.INFO, format="%(levelname)s | %(message)s")
logger = logging.getLogger(__name__)


def load_questions(xlsx_path):
    """从 xlsx 加载考核试题"""
    wb = openpyxl.load_workbook(xlsx_path)
    questions = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        qs = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            seq, question = row[0], row[1]
            if question and str(question).strip():
                qs.append({"seq": int(seq) if seq else len(qs) + 1, "question": str(question).strip()})
        questions[sheet_name] = qs
    return questions


def answer_questions(dxf_path, questions, llm_model="deepseek-v4-pro[1m]"):
    """对图纸运行 Agent 并回答所有问题"""
    # 使用独立的缓存文件，避免不同图纸之间串数据
    dxf_name = os.path.splitext(os.path.basename(dxf_path))[0]
    index_dir = os.path.join(os.path.dirname(dxf_path), "index")
    os.makedirs(index_dir, exist_ok=True)
    index_file = os.path.join(index_dir, f"{dxf_name}_parameters.json")

    # 删除已存在的缓存（确保从原始数据重建）
    if os.path.exists(index_file):
        os.remove(index_file)

    # 配置 Agent
    config = AgentConfig(
        cad_file=dxf_path,
        data_dir=os.path.dirname(dxf_path),
        index_file=index_file,
        llm_enabled=True,
        llm_provider="anthropic",
        llm_model=llm_model,
    )

    # 清除 session，避免多轮对话干扰
    agent = Agent(config)
    agent.load_cad(dxf_path)
    # 重置 session 历史
    agent._session.clear()
    logger.info(f"Agent ready: {agent.parameter_count} parameters indexed")

    # 逐个回答问题
    results = []
    for q in questions:
        try:
            answer = agent.ask(q["question"])
            logger.info(f'Q{q["seq"]}: {q["question"][:40]}... -> {answer[:60]}...')
        except Exception as e:
            answer = f"[ERROR] {e}"
            logger.error(f'Q{q["seq"]} failed: {e}')
        results.append({
            "seq": q["seq"],
            "question": q["question"],
            "answer": answer,
        })
    return results


def save_results(results, output_path):
    """保存答案"""
    # CSV 格式
    csv_path = output_path.replace(".json", ".csv")
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["序号", "问题", "答案"])
        for r in results:
            writer.writerow([r["seq"], r["question"], r["answer"]])
    logger.info(f"Saved CSV: {csv_path}")

    # JSON 格式
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    logger.info(f"Saved JSON: {output_path}")
    return csv_path


def main():
    import argparse

    parser = argparse.ArgumentParser(description="AstraDraft 最终考核答案生成")
    parser.add_argument("--questions", default="data/exam_temp/大赛CAD问题.xlsx",
                        help="考核试题xlsx路径")
    parser.add_argument("--output", default="data/exam_temp/answers",
                        help="输出路径前缀")
    parser.add_argument("--cad-dir", default="data/cad",
                        help="DXF文件目录")
    parser.add_argument("--model", default="deepseek-v4-pro[1m]",
                        help="LLM模型名称")
    args = parser.parse_args()

    # 加载试题
    questions = load_questions(args.questions)
    logger.info(f"Loaded questions: {', '.join(f'{k}({len(v)}题)' for k, v in questions.items())}")

    all_results = {}

    # 图纸映射：sheet名 -> DXF路径
    drawing_map = {
        "侧板": os.path.join(args.cad_dir, "侧板MODIFY.dxf"),
        "印刷版": os.path.join(args.cad_dir, "印刷件MODIFY.dxf"),
    }

    for sheet_name, qs in questions.items():
        # 找到对应的CAD文件
        dxf_key = sheet_name.replace("版", "")
        dxf_path = drawing_map.get(sheet_name) or drawing_map.get(dxf_key)

        if not dxf_path or not os.path.exists(dxf_path):
            logger.error(f"Cannot find CAD file for sheet '{sheet_name}'")
            continue

        logger.info(f"\n{'='*60}")
        logger.info(f"Processing: {sheet_name} ({dxf_path})")
        logger.info(f"Questions: {len(qs)}")

        results = answer_questions(dxf_path, qs, llm_model=args.model)
        all_results[sheet_name] = results

        # 打印结果
        print(f"\n{'='*60}")
        print(f"  {sheet_name} 答案")
        print(f"{'='*60}")
        for r in results:
            print(f"  Q{r['seq']:2d}. {r['question']}")
            print(f"      答: {r['answer']}")
            print()

    # 保存全部结果
    output_path = f"{args.output}.json"
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(all_results, f, ensure_ascii=False, indent=2)
    logger.info(f"Saved all results to {output_path}")

    # 保存每个sheet的CSV
    for sheet_name, results in all_results.items():
        csv_path = f"{args.output}_{sheet_name}.csv"
        with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["序号", "问题", "答案"])
            for r in results:
                writer.writerow([r["seq"], r["question"], r["answer"]])
        logger.info(f"Saved CSV: {csv_path}")

    print(f"\n{'='*60}")
    print(f"  全部完成！答案已保存到 {output_path}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
